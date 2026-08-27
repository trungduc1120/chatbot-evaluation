import argparse
import os
import sys

import pandas as pd
from openpyxl import load_workbook


MODEL_COLUMN_MAP = {
    # model: (respond_col_letter, elapsed_col_letter)
    "mistralai/ministral-8b": ("O", "Q"),
    "mistralai/mistral-7b-instruct": ("L", "N"),
    "meta-llama/llama-3.1-8b-instruct": ("I", "K"),
    "microsoft/phi-4": ("R", "T"),
}


def col_letter_to_index(letter: str) -> int:
    letter = letter.strip().upper()
    result = 0
    for ch in letter:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"Invalid column letter: {letter}")
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result  # 1-based index for openpyxl


def main() -> int:
    parser = argparse.ArgumentParser(description="Merge OpenRouter CSV results into Excel sheet")
    parser.add_argument("--excel_in", default="/home/vdsmart/eyepro-insight/tools/input_question.xlsx", help="Input Excel path")
    parser.add_argument("--csv", default="/home/vdsmart/eyepro-insight/openrouter_multimodel_logs.csv", help="CSV with columns: user_prompt, respond, model, elapsed_time")
    parser.add_argument("--excel_out", default="/home/vdsmart/eyepro-insight/tools/output_question.xlsx", help="Output Excel path")
    parser.add_argument("--sheet_index", type=int, default=1, help="Sheet index (0-based), default: 1 => sheet 2")
    parser.add_argument("--start_row", type=int, default=4, help="Start row (1-based), default: 4")
    args = parser.parse_args()

    if not os.path.exists(args.excel_in):
        print(f"Excel input not found: {args.excel_in}")
        return 1
    if not os.path.exists(args.csv):
        print(f"CSV not found: {args.csv}")
        return 1

    try:
        # Auto-detect delimiter
        df = pd.read_csv(args.csv, sep=None, engine="python")
    except Exception as e:
        print(f"Failed to read CSV: {e}")
        return 1

    # Ensure required columns exist (model may be absent if minimal CSV was created)
    for col in ["respond", "elapsed_time"]:
        if col not in df.columns:
            print(f"Missing column in CSV: {col}")
            return 1
    if "model" not in df.columns:
        print("Missing 'model' column in CSV; cannot map to target columns.")
        return 1

    # Group rows by model preserving order
    model_to_rows = {m: [] for m in MODEL_COLUMN_MAP.keys()}
    for _, row in df.iterrows():
        model = str(row.get("model", ""))
        if model in model_to_rows:
            model_to_rows[model].append((str(row.get("respond", "")), str(row.get("elapsed_time", ""))))

    # Load workbook and target sheet
    wb = load_workbook(args.excel_in)
    try:
        ws = wb.worksheets[args.sheet_index]
    except Exception:
        print(f"Invalid sheet index: {args.sheet_index}")
        return 1

    # Write each model's results to specified columns starting from start_row
    for model, (resp_col, time_col) in MODEL_COLUMN_MAP.items():
        rows = model_to_rows.get(model, [])
        if not rows:
            continue
        resp_col_idx = col_letter_to_index(resp_col)
        time_col_idx = col_letter_to_index(time_col)
        for i, (resp_text, elapsed_val) in enumerate(rows):
            r = args.start_row + i
            ws.cell(row=r, column=resp_col_idx).value = resp_text
            ws.cell(row=r, column=time_col_idx).value = elapsed_val

    # Save to new file
    try:
        wb.save(args.excel_out)
    except Exception as e:
        print(f"Failed to save Excel: {e}")
        return 1

    print(f"Wrote merged output to {args.excel_out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())


